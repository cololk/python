#-*- coding:utf-8 -*-
#本程式為調適SPEC COSE的資料庫建立片段
#適用python3.6版
import openpyxl
import multiprocessing

def loadEX(file):
	print("Program Starting...")
	wb=openpyxl.load_workbook(file)
	sheetname=wb.get_sheet_names()
	getsheet=wb.get_sheet_by_name(sheetname[0])
	print("載入%s..."% sheetname[0])	
	print(u"開始建立件號組成資料...")
	return getsheet, wb

def GereratePartContent(num,getsheet):	
	
	PartContent={}  # PartContent={'CM1':['AA1J','AA1B'], 'PD1':['AA1R','AB1A',...],...}
	for j in range(5, getsheet.max_row+1): # j是圖譜的列數
	        if getsheet.cell(row=j, column=getsheet.max_column-(16-num)).value==u'\u25cf':
	            for k in range(11, getsheet.max_column-17): # k是車規的欄數
	                if getsheet.cell(row=j, column=k).value==u'\u25cf':
	                    if getsheet.cell(row=j, column=7).value not in PartContent: # 如果此SPEC CODE是第一次在這個件號出現,就設定其值
	                        PartContent[getsheet.cell(row=j,column=7).value]=[getsheet.cell(row=4,column=k).value]
	                        #print("Add %s"% getsheet.cell(row=4,column=k).value )
	                    else: #否則就改為添加其值
	                        PartContent[getsheet.cell(row=j,column=7).value].append(getsheet.cell(row=4,column=k).value)
	                        #print("Add %s"% getsheet.cell(row=4,column=k).value )

	print(u"件號 %s SPEC CODE組成建立" % getsheet.cell(row=4, column=getsheet.max_column-(16-num)).value)
	return  getsheet.cell(row=4, column=getsheet.max_column-(16-num)).value, PartContent

#========進行sheet比對分析=================
def sheet_Analysis(wb2, sheetname2,lis ):
	analysis_result={}
	starttime=time.time()	
	print("展開比對程序...".decode('utf8'))
	for n in range(11, wb2.get_sheet_by_name(sheetname2[0]).max_column):
		if wb2.get_sheet_by_name(sheetname2[0]).cell(row=3, column=n).value==u'件號':
			g=n+1 # g值代表要開始遍歷的起始點
			print(u"取得分析起始欄位%s" % g)
	for h in range(0,len(sheetname2)):  #取得載入的EXCEL共有幾個Sheet
	        get=wb2.get_sheet_by_name(sheetname2[h])
	        print("載入%s".decode('utf8')% sheetname2[h])
	        for c in range(len(lis)):
	        	if long(Reg.findall(sheetname2[h])[0])==lis[c][0]:
	        		for i in range (5, get.max_row+1):
	        			if get.cell(row=i, column=7).value in lis[c][1]: #如果SPEC_CODE有出現在lis[c][1]中(這邊是指24010)
	        				for j in range(len(lis[c][1][get.cell(row=i, column=7).value])): #找出SPEC_CODE含有有幾個車規,一一遍歷過
	        					for k in range(g, get.max_column+1):
	        						if get.cell(row=4, column=k).value==lis[c][1][get.cell(row=i, column=7).value][j]:
	        							if lis[c][0] not in analysis_result:       #如果該件號的字典還沒建立
	        								analysis_result[lis[c][0]]=[[i,k]]     #就創建其值
	        							else:
	        								analysis_result[lis[c][0]].append([i,k]) #否則就改為添加值


#=========載入sheet_all檔案===========
def loadEX_sheet(file2):
	print(u"開始載入分頁檔案")
	print(u"載入 %s..." % file2)
	wb2=openpyxl.load_workbook(file2)
	sheetname2=wb2.get_sheet_names()
	print(u"分頁檔案共有 %s 個sheet" % len(sheetname2))
	for i in range(len(sheetname2)):
		print("[%s] %s" % (i+1,sheetname2[i]))
	print(u"載入完成")
	return wb2,sheetname2


if __name__=='__main__':
        results=[] #此為多執行緒物件返回值串列
        res=[]  #此為用get()方法翻譯後的串列,[['24010',{'CM1':['AA1J','AA1B']}], ['24011':{'CM1':['AA1C','AA1G']}]]
        file='SPEC_CODE.xlsx'
        file2='sheet_all.xlsx'
        getsheet,wb=loadEX(file)
        pool=multiprocessing.Pool()
        for num in range(17):
                result=pool.apply_async(GereratePartContent, (num,getsheet))
                results.append(result)
        pool.close()
        pool.join()
        print("The length of results is %d" % len(results))
        for i in range(17):
        	res.append(results[i].get())
        print(u"建立完成")
        print(res[0][0])

        loadEX_sheet(file2)









