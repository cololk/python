#-*- coding:utf-8 -*-
#本程式適用python3.4
#移除視窗功能、移除多線程功能

import sys
import os
import openpyxl
from openpyxl.styles import PatternFill
import re
import time #使用計時模組
import multiprocessing as mp

Color_Green=PatternFill(fgColor='008000', fill_type="solid")
Color_Red=PatternFill(fgColor='ff0000', fill_type="solid")
Reg=re.compile(r'\d\d\d\d\d')
lis=[] # lis的組成內容=[24010,SPEC_CODE,24011,SPEC_CODE,24012,SPEC_CODE]

def spec_code_analysis():
	file='SPEC_CODE.xlsx'
	#載入圖譜總表Excel
	print("Program Starting...")
	print(u"載入 %s..." % os.path.basename(file))
	wb=openpyxl.load_workbook(file)
	ws=wb.worksheets[0]
	print(u"載入%s..." % ws.title)	

	#件號串列資料庫建立
	print(u"開始建立件號組成資料...")
	def GereratePartContent(num,ws):
		print(u"件號 %s 組成設定展開..." % ws.cell(row=4, column=ws.max_column-(16-num)).value)
		PartContent={}  # PartContent={'CM1':['AA1J','AA1B'], 'PD1':['AA1R','AB1A',...],...}
		for j in range(5, ws.max_row+1): # j是圖譜的列數
		        if ws.cell(row=j, column=ws.max_column-(16-num)).value==u'\u25cf':
		            for k in range(11, ws.max_column-17): # k是車規的欄數
		                if ws.cell(row=j, column=k).value==u'\u25cf':
		                    if ws.cell(row=j, column=7).value not in PartContent: # 如果此SPEC CODE是第一次在這個件號出現,就設定其值
		                        PartContent[ws.cell(row=j,column=7).value]=[ws.cell(row=4,column=k).value]
		                        #print("Add %s"% ws.cell(row=4,column=k).value )
		                    else: #否則就改為添加其值
		                        PartContent[ws.cell(row=j,column=7).value].append(ws.cell(row=4,column=k).value)
		                        #print("Add %s"% ws.cell(row=4,column=k).value )
		return  ws.cell(row=4, column=ws.max_column-(16-num)).value, PartContent
	pool=mp.Pool()
	result=[]
	for i in range(17): #range(i)此處控制哪幾個件號需納入分析
		res=pool.apply_async(GereratePartContent, (i,ws))
		result.append(res)
	pool.close()
	pool.join()
	results=[]
	for item in result:
		results.append(item.get())
		print(item.get()[0])
	print(u"建立完成")


if __name__=='__main__':
	spec_code_analysis()


"""
#==============多執行緒part2======================
class loadExcelSheet_Thread(threading.Thread):
    def __init__(self, parent, value):
        threading.Thread.__init__(self)
        self._parent = parent
        self._value = value

    def run(self):
		#載入各分頁Excel
		global sheetname2, wb2
		print(u"開始載入分頁檔案")
		print(u"載入 %s..." % os.path.basename(file2))
		wb2=openpyxl.load_workbook(file2)
		sheetname2=wb2.get_sheet_names()
		print(u"分頁檔案共有 %s 個sheet" % len(sheetname2))
		for i in range(len(sheetname2)):
			print("[%s] %s" % (i+1,sheetname2[i]))
		print(u"載入完成")
		evt = CountEvent(myEVT_COUNT, -1, self._value)
		wx.PostEvent(self._parent, evt)

#=============多執行緒part3=================
class analysis_Thread(threading.Thread):
    def __init__(self, parent, value):

        threading.Thread.__init__(self)
        self._parent = parent
        self._value = value

    def run(self):
		#展開比對程序	
		starttime=time.time()	
		print("展開比對程序...".decode('utf8'))
		for n in range(11, wb2.get_sheet_by_name(sheetname2[0]).max_column):
			if wb2.get_sheet_by_name(sheetname2[0]).cell(row=3, column=n).value==u'件號':
				g=n+1
				print(u"取得分析起始欄位%s" % g)
		for h in range(0,len(sheetname2)):  #取得載入的EXCEL共有幾個Sheet
		        get=wb2.get_sheet_by_name(sheetname2[h])
		        print("載入%s".decode('utf8')% sheetname2[h])
		        if long(Reg.findall(sheetname2[h])[0]) in lis:  #須轉換為long型態才能跟sheetname2的頁面名稱匹配
		        	c=lis.index(long(Reg.findall(sheetname2[h])[0]))
			        d=c+1 # c為件號對應的索引位置,而其資料位於c+1處
			        print("取得 %s 對應圖譜總表索引位置 %s".decode('utf8') % (sheetname2[h],d))
			        for i in range (5, get.max_row+1):
			                if get.cell(row=i, column=7).value in lis[d]: #如果SPEC_CODE有出現在lis[1]中(這邊是指24010)
			                        for j in range(len(lis[d][get.cell(row=i, column=7).value])): #找出該件號該Spec_Code共出現幾個車規			
			                                for k in range(g, get.max_column+1):        
													# 如果依序遍歷車規發現有與該件號該SPE_CODE相同的就打上記號並標註為綠色
			                                        if get.cell(row=4, column=k).value==lis[d][get.cell(row=i, column=7).value][j]:
			                                                if get.cell(row=i, column=k).value is not None: #如果該欄位不是空,那就注入綠色(代表該選有選)
			                                                        get.cell(row=i, column=k).value=u'\u25cf'
			                                                        get.cell(row=i, column=k).fill=Color_Green
			                                                else: #否則該欄位就是空,那就注入紅色(代表該選沒選)
			                                                        get.cell(row=i, column=k).value=u'\u25cf'
			                                                        get.cell(row=i, column=k).fill=Color_Red

		path, filename = os.path.split(file2)
		saveFileName=path+'\\OUTPUT_'+filename                                          
		wb2.save(saveFileName)
		print(u"檔案產出: %s" % os.path.basename(saveFileName))
		endtime=time.time()
		total=endtime-starttime
		print "used time:", total
		print("Done")
		evt = CountEvent(myEVT_COUNT, -1, self._value)
		wx.PostEvent(self._parent, evt)


if __name__=='__main__':
	app=wx.App()
	main_win=Myframe(None)
	main_win.init_Myframe()
	main_win.init_redirectText()
	main_win.Show()
	app.MainLoop()
"""

        
    
