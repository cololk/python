#-*- coding:utf-8 -*-
#本程式可進行圖譜的自動分析

import openpyxl
#import sys  #python2.7會出現編碼錯誤之類的問題,google說要加入這三行
#reload(sys) #python2.7會出現編碼錯誤之類的問題,google說要加入這三行
#sys.setdefaultencoding('utf-8') #python2.7會出現編碼錯誤之類的問題,google說要加入這三行

print("Program Starting...")
wb=openpyxl.load_workbook('GPK_DRAWING_LIST_AM_test_test.xlsx')
sheetname=wb.get_sheet_names()
getsheet=wb.get_sheet_by_name(sheetname[0])
print("載入%s..." % sheetname[0])

# column=7: SPEC CODE位置
# row=4: 車規位置
def GereratePartContent(num):
	PartContent={}
	for j in range(5, getsheet.max_row+1): # j是圖譜的列數
	        if getsheet.cell(row=j, column=getsheet.max_column-(16-num)).value==u'\u25cf':
	            for k in range(11, getsheet.max_column-17): # k是車規的欄數
	                if getsheet.cell(row=j, column=k).value==u'\u25cf':
	                    if getsheet.cell(row=j, column=7).value not in PartContent: # 如果此SPEC CODE是第一次在這個件號出現,就設定其值
	                        PartContent[getsheet.cell(row=j,column=7).value]=[getsheet.cell(row=4,column=k).value]
	                        print("Add %s"% getsheet.cell(row=4,column=k).value )
	                    else: #否則就改為添加其值
	                        PartContent[getsheet.cell(row=j,column=7).value].append(getsheet.cell(row=4,column=k).value)
	                        print("Add %s"% getsheet.cell(row=4,column=k).value )
	return  getsheet.cell(row=4, column=getsheet.max_column-(16-num)).value, PartContent

lis=[]
for i in range(1): #range(i)此處控制哪幾個件號需納入分析
	lis.append(GereratePartContent(i)[0])
	lis.append(GereratePartContent(i)[1])
	print("件號序列 %s SPEC CODE組成內容為:%s" % (GereratePartContent(i)[0],GereratePartContent(i)[1]))

        
    
