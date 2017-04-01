#-*- coding:utf-8 -*-
#本程式可進行圖譜的自動分析
import wx
import sys
import openpyxl
from openpyxl.styles import PatternFill
import frame
import re
Color_Green=PatternFill(fgColor='008000', fill_type="solid")
Color_Red=PatternFill(fgColor='ff0000', fill_type="solid")
Reg=re.compile(r'\d\d\d\d\d')
lis=[] # lis的組成內容=[24010,SPEC_CODE,24011,SPEC_CODE,24012,SPEC_CODE]

#=================載入Excel================
print("Program Starting...")
wb=openpyxl.load_workbook('GPK_DRAWING_LIST_AM_For_Test.xlsx')
sheetname=wb.get_sheet_names()
getsheet=wb.get_sheet_by_name(sheetname[0])
print("載入%s...".decode('utf8') % sheetname[0])

#===============視窗介面框架================
#object要代入欲被轉換為Stdin輸出容器的TextCtrl物件
class RedirectText(object):
    def __init__(self,aWxTextCtrl):
        self.out=aWxTextCtrl

    def write(self,string):
        self.out.WriteText(string)

class Myframe(frame.MyFrame1):
	def init_Myframe(self):
		self.m_textCtrl1.SetValue("GPK_DRAWING_LIST_AM_For_Test.xlsx")

	def init_redirectText(self):  #這是自行定義的方法,在後面的主函數中要特別呼叫
		#輸出重新導向
		redir=RedirectText(self.m_textCtrl2)
		sys.stdout=redir

	def loadExcel(self,event):
		# column=7: SPEC CODE位置
		# row=4: 車規位置
		#==============件號串列資料庫建立======================
		print(u"開始建立件號組成資料...")
		def GereratePartContent(num):
			PartContent={}  # PartContent={'CM1':['AA1J','AA1B'], 'PD1':['AA1R','AB1A',...],...}
			for j in range(5, getsheet.max_row+1): # j是圖譜的列數
			        if getsheet.cell(row=j, column=getsheet.max_column-(16-num)).value==u'\u25cf':
			            for k in range(11, getsheet.max_column-17): # k是車規的欄數
			                if getsheet.cell(row=j, column=k).value==u'\u25cf':
			                    if getsheet.cell(row=j, column=7).value not in PartContent: # 如果此SPEC CODE是第一次在這個件號出現,就設定其值
			                        PartContent[getsheet.cell(row=j,column=7).value]=[getsheet.cell(row=4,column=k).value]
			                        #print("Add %s"% getsheet.cell(row=4,column=k).value )
			                    else: #否則就改為添加其值
			                        PartContent[getsheet.cell(row=j,column=7).value].append(getsheet.cell(row=4,column=k).value)
			                        #print("Add %s"% getsheet.cell(row=4,column=k).value )
			return  getsheet.cell(row=4, column=getsheet.max_column-(16-num)).value, PartContent
		
		for i in range(17): #range(i)此處控制哪幾個件號需納入分析
			lis.append(GereratePartContent(i)[0])
			lis.append(GereratePartContent(i)[1])
			print("件號 %s SPEC CODE組成建立".decode('utf8') % GereratePartContent(i)[0])
		print(u"建立完成")

	def analysis(self,event):
		#=================================展開比對程序===============================
		print("展開比對程序...".decode('utf8'))
		for h in range(1,len(sheetname)):
		        get=wb.get_sheet_by_name(sheetname[h])
		        print("載入%s".decode('utf8')% sheetname[h])
		        if long(Reg.findall(sheetname[h])[0]) in lis:
		        	c=lis.index(long(Reg.findall(sheetname[h])[0]))
			        d=c+1 # c為件號對應的索引位置,而其資料位於c+1處
			        print("取得 %s 對應圖譜總表索引位置 %s".decode('utf8') % (sheetname[h],d))
			        for i in range (5, get.max_row+1):
			                if get.cell(row=i, column=7).value in lis[d]: #如果SPEC_CODE有出現在lis[1]中(這邊是指24010)
			                        for j in range(len(lis[d][get.cell(row=i, column=7).value])): #找出該件號該Spec_Code共出現幾個車規
			                                for k in range(11, get.max_column+1):
			                                        
													# 如果依序遍歷車規發現有與該件號該SPE_CODE相同的就打上記號並標註為綠色
			                                        if get.cell(row=4, column=k).value==lis[d][get.cell(row=i, column=7).value][j]:
			                                                if get.cell(row=i, column=k).value is not None: #如果該欄位不是空,那就注入綠色(代表該選有選)
			                                                        get.cell(row=i, column=k).value=u'\u25cf'
			                                                        get.cell(row=i, column=k).fill=Color_Green
			                                                else: #否則該欄位就是空,那就注入紅色(代表該選沒選)
			                                                        get.cell(row=i, column=k).value=u'\u25cf'
			                                                        get.cell(row=i, column=k).fill=Color_Red
		                                                

		wb.save('GPK_DRAWING_LIST_AM_OUTPUT.xlsx')
		print("Save File to GPK_DRAWING_LIST_AM_OUTPUT.xlsx")
		print("Done")

if __name__=='__main__':
	app=wx.App()
	main_win=Myframe(None)
	main_win.init_Myframe()
	main_win.init_redirectText()
	main_win.Show()
	app.MainLoop()


        
    
