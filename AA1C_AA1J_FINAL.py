#-*- coding:utf-8 -*-
#本程式可比對兩個件件號之間SPEC CODE的差異

import openpyxl

def trans(cgetsheet):
	lis={}
	for i in range(3,cgetsheet.max_row+1):
		lis[cgetsheet.cell(row=i,column=1).value]=cgetsheet.cell(row=i, column=2).value
	return lis
    
<<<<<<< HEAD
wb=openpyxl.load_workbook('C:\\demo\\python\\B版改C版差異_AH_AK.xlsx')
=======
wb=openpyxl.load_workbook('C:\\Python27\\TEST002.xlsx')
>>>>>>> ea2327774cd354ebc46a4e98a8fa53b083ddc24b
sheetname=wb.get_sheet_names()
for i in range(0,len(sheetname)):
    getsheet=wb.get_sheet_by_name(sheetname[i])
    if i==0:
    	L1=trans(getsheet)
    	s1=set(L1)
    else:
    	L2=trans(getsheet)
    	s2=set(L2)
print("第一組差異為:")
print(s1.symmetric_difference(s2))




