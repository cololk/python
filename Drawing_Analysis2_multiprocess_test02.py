#-*- coding:utf-8 -*-
#本程式適用python3.4
#移除視窗功能、移除多線程功能，移除queue進行列隊處理的功能
#進度:須進行writer()內容的定義,尚未完成
'''
紀錄:
1.發現使用multiprocess處理openpyxl會導致記憶體內容被完整複製,將爆量使用記憶體,
公司電腦記憶體僅有2Gb,無法使用此程式。
2.使用queue進行列隊處理時,在最後儲存檔案時會報錯,但使用簡單版的excel進行測試卻無問題,
可能是sheet_all.xlsx有某種未知參數導致此問題,故本程式test02版先移除queue列隊處理功能
'''
import sys
import os
import openpyxl
import queue
from openpyxl.styles import PatternFill
import re
import time #使用計時模組
import multiprocessing as mp

Color_Green=PatternFill(fgColor='008000', fill_type="solid")
Color_Red=PatternFill(fgColor='ff0000', fill_type="solid")
Reg=re.compile(r'\d\d\d\d\d')
lis=[] # lis的組成內容=[24010,SPEC_CODE,24011,SPEC_CODE,24012,SPEC_CODE]
class wsDiff(object):
    def __init__(self, row, column, value):
        self.row = row
        self.column = column
        self.value = value

def GereratePartContent(num,ws): #主程序1的副程式,主要進行多進程處理
	print(u"件號 %s 組成設定展開..." % ws.cell(row=4, column=ws.max_column-(16-num)).value)		
	PartContent={}  # PartContent={'CM1':['AA1J','AA1B'], 'PD1':['AA1R','AB1A',...],...}
	for j in range(5, ws.max_row+1): # j是圖譜的列數
	        if ws.cell(row=j, column=ws.max_column-(16-num)).value==u'\u25cf':
	            for k in range(11, ws.max_column-17): # k是車規的欄數
	                if ws.cell(row=j, column=k).value==u'\u25cf':
	                    if ws.cell(row=j, column=7).value not in PartContent: # 如果此SPEC CODE是第一次在這個件號出現,就設定其值
	                        PartContent[ws.cell(row=j,column=7).value]=[ws.cell(row=4,column=k).value]
	                        #print("Add %s"% ws.cell(row=4,column=k).value )
	                    else: #否則就改為添加其值
	                        PartContent[ws.cell(row=j,column=7).value].append(ws.cell(row=4,column=k).value)
	                        #print("Add %s"% ws.cell(row=4,column=k).value )
	return  ws.cell(row=4, column=ws.max_column-(16-num)).value, PartContent

def spec_code_analysis(): #主程序1,負責分析圖譜總表
	file='c:\\Python34\\SPEC_CODE.xlsx'
	print(u"載入 %s..." % os.path.basename(file))
	wb=openpyxl.load_workbook(file)
	ws=wb.worksheets[0]
	print(u"載入%s..." % ws.title)
	
	pool=mp.Pool()
	result=[]
	for i in range(17): #range(i)此處控制哪幾個件號需納入分析
		res=pool.apply_async(GereratePartContent, (i,ws))
		result.append(res)
	pool.close()
	pool.join()
	results=[]
	for item in result:
		results.append(item.get())
	print(u"建立完成")
	return results

def ws_job(wb, ws_idx,results):主程序2副程式,主要進行多進程處理
	diff=[]
	ws = wb.worksheets[ws_idx]
	print('[PID %s]: process (%s)' % (os.getpid(), ws.title))
	for n in range(11, wb.worksheets[0].max_column):
		if wb.worksheets[0].cell(row=3, column=n).value==u'件號':
			g=n+1
			print(u"取得分析起始欄位%s" % g)
	print(u"載入%s"% ws.title)
	for res in results:
	    if str(Reg.findall(ws.title)[0]) == str(res[0]):  #須轉換為long型態才能跟sheetname2的頁面名稱匹配
	        print(u"取得 %s 對應圖譜總表項目: %s" % (ws.title,res[0]))
	        for i in range (5, ws.max_row+1):
	                if ws.cell(row=i, column=7).value in res[1]: #如果SPEC_CODE有出現在lis[1]中(這邊是指24010)
	                        for j in range(len(res[1][ws.cell(row=i, column=7).value])): #找出該件號該Spec_Code共出現幾個車規			
	                                for k in range(g, ws.max_column+1):        
											# 如果依序遍歷車規發現有與該件號該SPE_CODE相同的就打上記號並標註為綠色
	                                        if ws.cell(row=4, column=k).value==res[1][ws.cell(row=i, column=7).value][j]:
	                                                if ws.cell(row=i, column=k).value is not None: #如果該欄位不是空,那就注入綠色(代表該選有選)
	                                                	diff.append([i,k,u'\u25cf',Color_Green])
	                                                    #ws.cell(row=i, column=k).value=u'\u25cf'
	                                                    #ws.cell(row=i, column=k).fill=Color_Green
	                                                else: #否則該欄位就是空,那就注入紅色(代表該選沒選)
	                                                	diff.append([i,k,u'\u25cf',Color_Red])
	                                                    #ws.cell(row=i, column=k).value=u'\u25cf'
	                                                    #ws.cell(row=i, column=k).fill=Color_Red
	return ws.title, diff    

def writer(Diffs,wb):#主程序3:進行檔案資料寫入
	print("寫入 %s ..." % 'sheet_24010_24080.xlsx')
	wb3 = openpyxl.load_workbook('sheet_24010_24080.xlsx')
	for index in Diffs:
		#尚待完成

def sheet_analysis(results): #主程序2:負責比對sheet與SPEC_CODE資料,並返回座標串列
	print("載入 %s ..." % 'sheet_24010_24080.xlsx')
	wb2 = openpyxl.load_workbook('sheet_24010_24080.xlsx')

	pool=mp.Pool()
	Diff=[]
	for i in range(len(wb2.worksheets)):
		res=pool.apply_async(ws_job,(wb, i,results))
		Diff.append(res)
	pool.close()
	pool.join()
	Diffs=[]
	for res in Diff:
		Diffs.append(res.get())
	print('Done')
	return Diffs #格式為[('24010',[填值座標]) , ('24011',[填值座標]),...]


	
if __name__=='__main__':
	results=spec_code_analysis()
	Diffs=sheet_analysis(results)
	


      
    
