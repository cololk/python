#本程式可以讓使用者輸入欲查詢的料號,會返回資料庫比對確認的結果
# -*- coding: utf-8 -*-
import openpyxl
import re
from datetime import datetime
import tkinter as tk

def loadExcel(filename): #返回某個EXCEL檔案的所有sheet的串列
    wb=openpyxl.load_workbook(filename)
    sheetname=wb.get_sheet_names()
    return wb,sheetname

def theSameName(dit,key, getsheet,i, next_line,col_num,col_maker ):
    j=0
    while getsheet.cell(row=i+next_line+j, column=col_num).value is not None:
        if getsheet.cell(row=i+next_line+j, column=col_num).value.startswith('H:'):
            #比對正確就把接頭料號與廠商名稱加入該鍵的串列
            dit[key].append(getsheet.cell(row=i+next_line+j, column=col_num).value.lstrip('H:').strip())
            dit[key].append(getsheet.cell(row=i+next_line+j, column=col_maker).value)

            T13_theSame = getsheet.cell(row=i+next_line+j+1, column=col_num).value
            HT25_theSame = getsheet.cell(row=i+next_line+j+1, column=col_maker).value           #原廠端子廠商
            if T13_theSame is None:
                pass
            else:
                k=1
                while T13_theSame is not None:
                    if T13_theSame.startswith("T:"):
                        break
                    else:
                        T13_theSame = getsheet.cell(row=i+next_line+j+1+k, column=col_num).value             #原廠端子料號
                        HT25_theSame = getsheet.cell(row=i+next_line+j+1+k, column=col_maker).value           #原廠端子廠商
                    k += 1

            dit[key].append(T13_theSame)
            dit[key].append(HT25_theSame)
        else:
            pass
        j +=1
    return dit

# cscc()的格式為cscc(字典檔,i, Excel_sheet,功能名稱的欄號, 接頭料號的欄號,製造商的欄號)
def cscc(dit,i, getsheet,col_function,col_num,col_maker ):
    
    if getsheet.cell(row=i, column=col_function).value is None: #column=12欄位是功能名稱
        pass
    else:
        if getsheet.cell(row=i, column=col_function).value.upper().strip().replace('_',' ').replace('-',' ') not in dit: #如果功能名稱未出現在CSSS_CONN_NAME中,此功能名稱鍵還未被創立
            #將字串轉換成大寫,移除前後換行空白字元,底線換成空格,中線換成空格
            L12=getsheet.cell(row=i, column=col_function).value.upper().strip().replace('_',' ').replace('-',' ')                        
            H13 = getsheet.cell(row=i+1, column=col_num).value                   # 原廠膠盒料號
            # 如果H13是空字符,那就將H13由原本指定的列數再往下一列,其餘元素跟著往下偏移
            if H13 is None:                                                                        
                H13  = getsheet.cell(row=i+2, column=col_num).value.lstrip('H:')  #原廠膠盒料號
                HF25 = getsheet.cell(row=i+2, column=col_maker).value             #原廠膠盒廠商
                T13  = getsheet.cell(row=i+3, column=col_num).value               #原廠端子料號
                HT25 = getsheet.cell(row=i+3, column=col_maker).value             #原廠端子廠商
                if T13 is None:
                    pass
                else:
                    j=1 #while迴圈迭代器
                    #如果T13不是以T:開頭,就再往下一行搜尋
                    while T13 is not None:
                        if T13.startswith("T:"):
                            break
                        else:
                            T13  = getsheet.cell(row=i+3+j, column=col_num).value             #原廠端子料號
                            HT25 = getsheet.cell(row=i+3+j, column=col_maker).value           #原廠端子廠商
                        j += 1
                    
                # cscc_conn_name[接頭名稱]={ 膠盒料號, 廠商, 端子料號,廠商} 
                dit[L12]=[H13,HF25,T13,HT25]
                #如果原行數下一行的值非"None",比對其是否以H:開頭,如果是就加入list中
                theSameName(dit,L12, getsheet,i,4,col_num,col_maker)
                                                    
            else:  #如果H13不是None, 就不用往下偏移
                HF25 = getsheet.cell(row=i+1, column=col_maker).value             # 原廠膠盒廠商
                T13  = getsheet.cell(row=i+2, column=col_num).value               #原廠端子料號
                HT25 = getsheet.cell(row=i+2, column=col_maker).value             #原廠端子廠商
                if T13 is None:
                    pass
                else:
                    j=1 #while迴圈迭代器
                    while T13 is not None:
                        if T13.startswith("T:"):
                            break
                        else:
                            T13  = getsheet.cell(row=i+2+j, column=col_num).value             #原廠端子料號
                            HT25 = getsheet.cell(row=i+2+j, column=col_maker).value           #原廠端子廠商
                        j += 1                
                # cscc_conn_name[接頭名稱]={ 膠盒料號, 廠商, 端子料號,廠商} 
                dit[L12]= [H13.lstrip('H:'),HF25,T13,HT25]
                theSameName(dit, L12, getsheet, i, 2,col_num,col_maker)

        else: # 功能名稱已在CSCC字典中,改為添加值
            L12=getsheet.cell(row=i, column=col_function).value.upper().strip().replace('_',' ').replace('-',' ')
            H13 = getsheet.cell(row=i + 1, column=col_num).value  # 原廠膠盒料號
            # 如果H13是空字符,那就將H13由原本指定的列數再往下一列,其餘元素跟著往下偏移
            if H13 is None:
                H13 = getsheet.cell(row=i + 2, column=col_num).value.lstrip('H:')  #原廠膠盒料號
                HF25 = getsheet.cell(row=i + 2, column=col_maker).value            #原廠膠盒廠商
                T13  = getsheet.cell(row=i+3, column=col_num).value                #原廠端子料號
                HT25 = getsheet.cell(row=i+3, column=col_maker).value              #原廠端子廠商
                if T13 is None:
                    pass
                else:
                    j=1 #while迴圈迭代器
                    #如果T13不是以T:開頭,且不是None,就再往下一行搜尋
                    while T13 is not None:
                        if T13.startswith("T:"):
                            break
                        else:
                            T13  = getsheet.cell(row=i+3+j, column=col_num).value             #原廠端子料號
                            HT25 = getsheet.cell(row=i+3+j, column=col_maker).value           #原廠端子廠商
                        j += 1
                # cscc_conn_name[接頭名稱]={ 膠盒料號, 廠商, 端子料號,廠商}
                dit[L12].append(H13)
                dit[L12].append(HF25)
                dit[L12].append(T13)
                dit[L12].append(HT25)
                print("The connector %s append %s" % (L12, (HF25+' '+H13)))         
                # 如果原行數下一行的值非"None",比對其是否以H:開頭,如果是就加入list中
                theSameName(dit, L12, getsheet, i, 3,col_num,col_maker)


            else:  # 如果H13不是None, 就不用往下偏移
                HF25 = getsheet.cell(row=i + 1, column=col_maker).value  # 原廠膠盒廠商
                T13  = getsheet.cell(row=i+2, column=col_num).value               #原廠端子料號
                HT25 = getsheet.cell(row=i+2, column=col_maker).value             #原廠端子廠商
                if T13 is None:
                    pass
                else:
                    j=1 #while迴圈迭代器
                    #如果T13不是以T:開頭,且不是None,就再往下一行搜尋
                    while T13 is not None:
                        if T13.startswith("T:"):
                            break
                        else:
                            T13  = getsheet.cell(row=i+2+j, column=col_num).value             #原廠端子料號
                            HT25 = getsheet.cell(row=i+2+j, column=col_maker).value           #原廠端子廠商
                        j += 1                                
                # cscc_conn_name[接頭名稱]={ 膠盒料號, 廠商, 端子料號,廠商}
                dit[L12].append(H13.lstrip('H:'))
                dit[L12].append(HF25)
                dit[L12].append(T13)
                dit[L12].append(HT25)
                print("The connector %s append %s" % (L12, (HF25+' '+H13)))
                theSameName(dit, L12, getsheet, i, 2,col_num,col_maker)  
    return dit

#===================主要判斷邏輯===================
def finder():
    connector_no=enter_string.get()
    flag=0
    for csccname in cscc_conn:
        for j in range(0,len(cscc_conn[csccname]),4):
            if (cscc_conn[csccname][j+1]+' '+cscc_conn[csccname][j]).upper().strip().replace('_', ' ').replace('-',' ')==connector_no.upper().strip().replace('_', ' ').replace('-',' '):
                print("[+]Find %s is used on %s" % (connector_no, csccname))
                text1.insert('end', "[+]Find %s is used on %s" % (connector_no, csccname))
                flag=1
                break
    if flag==0:
        print("[-]Not find")
        text1.insert('end',"[-]Not find" )


#================載入樂榮CSCC ============
Data_local="c:\\Python34\\DATA\\"
CSCC_Name=['24010-KN711-CSCC-161214.xlsx', '24012-KN711-CSCC-161214.xlsx',
           '24068-KN711-CSCC-161214.xlsx','24023-KN711-CSCC-170307.xlsx']
print("創建Excel物件....CSCC")
cscc_conn={}
for CsccName in CSCC_Name:
    print("載入 %s" % (Data_local+CsccName))
    wb, sheetname=loadExcel(Data_local+CsccName)
    for shet in sheetname:
        getsheet=wb.get_sheet_by_name(shet)
        for i in range(25, int(getsheet.max_row)+1):
            cscc_conn=cscc(cscc_conn, i, getsheet,12,13,25)

#================載入矢崎CSCC==============
CsccName1='24011-KN711-CSCC-161104.xlsx'
print("載入 %s ..." % (Data_local+CsccName1))
wb3,sheetname3=loadExcel(Data_local+CsccName1)
getsheet3=wb3.get_sheet_by_name(sheetname3[0])
for i in range(25, int(getsheet3.max_row)+1):
    cscc_conn=cscc(cscc_conn,i,getsheet3,14,15,27)

#===============視窗程序================
window = tk.Tk()
window.title('CSCC_Finder')
window.geometry('300x200') #寬X高

enter_string = tk.Entry(window, show=None) #show選擇None,不會特別影藏輸入的字串
enter_string.pack()

button1 = tk.Button(window, text='Search', width=7,height=1, command=finder)
button1.pack()

text1 = tk.Text(window, height=150,width=190)
text1.pack()

window.mainloop()