#-*- coding:utf-8 -*-
#本程式可用來找出廠商料號對應的原廠接頭編號以及其對應的功能
import openpyxl
import re
import time
import wx
import baseFrame01 #記得要先IMPORT GUI介面要用的框架
import sys  #python2.7會出現編碼錯誤之類的問題,google說要加入這三行
reload(sys) #python2.7會出現編碼錯誤之類的問題,google說要加入這三行
sys.setdefaultencoding('utf-8') #python2.7會出現編碼錯誤之類的問題,google說要加入這三行


#=========================載入材規資料=========================
Data_local="c:\\Python34\\DATA\\"
CSCC_Name=['24014-LB110-CSCC--C-170213.xlsx','24015-LB110-CSCC--C-161031.xlsx',
                           '24068-LB110-CSCC--C-160923.xlsx','24093-LB110-CSCC--0-161031.xlsx']
print("創建Excel物件...載入CSCC")
cscc_no={}
for CsccName in CSCC_Name:
    print("載入 %s" % (Data_local+CsccName))
    wb = openpyxl.load_workbook(Data_local+CsccName)
    sheetname = wb.get_sheet_names()
    getsheet = wb.get_sheet_by_name(sheetname[0])

    #創建材規的附件與料號對照字典   
    for i in range(27, int(getsheet.max_row)+1):           #因24014的材規從27列才開始是資料,上限需+1才能讀取到最後一列本身
        DN118 = getsheet.cell(row=i, column=118).value     #DN欄位廠商料號轉換為數字為118
        DN118_str=str(DN118).strip()                       #全部轉換成字串型式,且去除頭尾字符
        L12 = getsheet.cell(row=i, column=12).value
        if L12 is None:
            pass
        else:
            L12strip=L12.lstrip('H:').lstrip('T:').strip()            #去除字頭的H:或者T: , 最後去除前後空白
        if DN118 is None:                                             #花了一堆時間確認noneType型式的判斷法...= =
            pass
        else:
            DN118strip=DN118_str.lstrip("附件")                 #去除字串頭的"附件"兩個字        
            #字典檔開始增加資料,格式為cscc_no[A]=[ B, C, D] 字典'鍵'對應的'值'為串列型式,        
            #cscc_no[A][0]為料件原廠編號, cscc_no[A][1]為料件廠商, cscc_no[A][2]為接頭功能名稱  ,A則為廠商料號,例如4312876         
            L12_1=getsheet.cell(row=i-1, column=12).value
            if L12_1.startswith('T:'):
                L12_1=getsheet.cell(row=i, column=12).value.strip()
            if L12_1.startswith('H:') or L12_1.startswith('H：'):
                L12_1=L12_1+'膠盒的端子'
            cscc_no[DN118strip] = [L12strip,  getsheet.cell(row=i, column=45).value, L12_1]  


#============將Stdin轉到StaticText的特殊處理方式===============
#object要代入欲被轉換為Stdin輸出容器的TextCtrl物件
class RedirectText(object):
    def __init__(self,aWxTextCtrl):
        self.out=aWxTextCtrl

    def write(self,string):
        self.out.WriteText(string)

#=================視窗程序介面===============
class Myframe(baseFrame01.baseWind):
	def init_Myframe(self):
		self.m_textCtrl1.SetValue("Input No")
	def Search(self,event):
		pass

	def init_redirectText(self):  #這是自行定義的方法,在後面的主函數中要特別呼叫
		#展開輸出重新導向
		redir=RedirectText(self.m_textCtrl2)
		sys.stdout=redir

	def Search(self,event):
		# ========================載入待比對資料EXCEL表=======================
		Data_local="c:\\Python34\\DATA\\"
		ob = openpyxl.load_workbook(Data_local+'LCH-DSHEET-170114.xlsx')
		obsheetname = ob.get_sheet_names()
		getobsheet=ob.get_sheet_by_name(obsheetname[1])

		#創建待比對資料串列
		object_no=[]
		for j in range(3,int(getobsheet.max_row)+1):
		    #增加進去的串列資料要加上str()方法,以確定都是以字串的形式加入,不能是數字,且去除頭尾字符
		    object_no.append(str(getobsheet.cell(row=j, column=2).value).strip()) 


		#======================找出待比對資料對應的料號((主程序))==============      
		print("Lunch......")
		time.sleep(1)
		find=0  #找出的資料計數器初始值
		# 找出的待比對廠商編號對應的原廠料號
		final_list={}  
		for attatch_no in object_no:
		    flag=0
		    for data in cscc_no:
		        if data==attatch_no:
		            print("廠商料號 %s 於CSCC中對應的原廠料號為 %s %s,其功能為%s" % (attatch_no, cscc_no[data][1],  cscc_no[data][0],cscc_no[data][2].strip()))        
		            flag=1        #找到資料,旗標就設定為1 
		            find=find+1
		            # 產生待比對附件編號對照料號表
		            final_list[attatch_no]=[cscc_no[data][0], cscc_no[data][1]]
		        else:
		            pass
		    if flag==0:
		        print("廠商料號 %s 無對應的原廠料號"% attatch_no)


		print("比對完畢")
		print("共有 %s 筆資料, 比對出 %s 項" % (len(object_no),find))
		print(" ")



if __name__=='__main__':
	app=wx.App()
	main_win=Myframe(None)
	main_win.init_Myframe()
	main_win.init_redirectText()
	main_win.Show()
	app.MainLoop()