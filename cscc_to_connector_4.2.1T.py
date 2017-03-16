#本程式可實現將GPK_MC CSCC中的料號填入Connector list
#且對於同功能名稱但有不同接頭的情況進行比對區分
#模糊比對採用find()函式手法, 缺點:無法比對同功能名稱,但字串順序不同的狀況
#模糊比對增加set()函式手法, 可比對同功能名稱但字串順序不同的狀況
#增加端子比對功能,包含樂榮與矢崎,但不含台裕線束
#4.2.1T版本追加: 如果端子廠商比對結果不同,就把儲存格反紅

import openpyxl
import re
from datetime import datetime
from openpyxl.styles import PatternFill
Color_red=PatternFill(fgColor='DC143C', fill_type="darkUp") #作用為如果廠商不同廠,可把儲存格顏色反紅
num=0
def loadExcel(filename): #返回某個EXCEL檔案的所有sheet的串列
    wb=openpyxl.load_workbook(filename)
    sheetname=wb.get_sheet_names()
    return wb,sheetname

def theSameName(dit,key, getsheet,i, next_line,col_num,col_maker ):
    j=0
    while getsheet.cell(row=i+next_line+j, column=col_num).value is not None:
        if getsheet.cell(row=i+next_line+j, column=col_num).value.startswith('H:'):
            #比對正確就把接頭料號與廠商名稱加入該鍵的串列
            dit[key].append(getsheet.cell(row=i+next_line+j, column=col_num).value.lstrip('H:').strip())
            dit[key].append(getsheet.cell(row=i+next_line+j, column=col_maker).value)

            T13_theSame = getsheet.cell(row=i+next_line+j+1, column=col_num).value
            HT25_theSame = getsheet.cell(row=i+next_line+j+1, column=col_maker).value           #原廠端子廠商
            if T13_theSame is None:
                pass
            else:
                k=1
                while T13_theSame is not None:
                    if T13_theSame.startswith("T:"):
                        break
                    else:
                        T13_theSame = getsheet.cell(row=i+next_line+j+1+k, column=col_num).value             #原廠端子料號
                        HT25_theSame = getsheet.cell(row=i+next_line+j+1+k, column=col_maker).value           #原廠端子廠商
                    k += 1

            dit[key].append(T13_theSame)
            dit[key].append(HT25_theSame)
        else:
            pass
        j +=1
    return dit

# cscc()的格式為cscc(字典檔,i, Excel_sheet,功能名稱的欄號, 接頭料號的欄號,製造商的欄號)
def cscc(dit,i, getsheet,col_function,col_num,col_maker ):
    
    if getsheet.cell(row=i, column=col_function).value is None: #column=12欄位是功能名稱
        pass
    else:
        if getsheet.cell(row=i, column=col_function).value.upper().strip().replace('_',' ').replace('-',' ') not in dit: #如果功能名稱未出現在CSSS_CONN_NAME中,此功能名稱鍵還未被創立
            #將字串轉換成大寫,移除前後換行空白字元,底線換成空格,中線換成空格
            L12=getsheet.cell(row=i, column=col_function).value.upper().strip().replace('_',' ').replace('-',' ')                        
            H13 = getsheet.cell(row=i+1, column=col_num).value                   # 原廠膠盒料號
            # 如果H13是空字符,那就將H13由原本指定的列數再往下一列,其餘元素跟著往下偏移
            if H13 is None:                                                                        
                H13  = getsheet.cell(row=i+2, column=col_num).value.lstrip('H:')  #原廠膠盒料號
                HF25 = getsheet.cell(row=i+2, column=col_maker).value             #原廠膠盒廠商
                T13  = getsheet.cell(row=i+3, column=col_num).value               #原廠端子料號
                HT25 = getsheet.cell(row=i+3, column=col_maker).value             #原廠端子廠商
                if T13 is None:
                    pass
                else:
                    j=1 #while迴圈迭代器
                    #如果T13不是以T:開頭,就再往下一行搜尋
                    while T13 is not None:
                        if T13.startswith("T:"):
                            break
                        else:
                            T13  = getsheet.cell(row=i+3+j, column=col_num).value             #原廠端子料號
                            HT25 = getsheet.cell(row=i+3+j, column=col_maker).value           #原廠端子廠商
                        j += 1
                    
                # cscc_conn_name[接頭名稱]={ 膠盒料號, 廠商, 端子料號,廠商} 
                dit[L12]=[H13,HF25,T13,HT25]
                #如果原行數下一行的值非"None",比對其是否以H:開頭,如果是就加入list中
                theSameName(dit,L12, getsheet,i,4,col_num,col_maker)
                                                    
            else:  #如果H13不是None, 就不用往下偏移
                HF25 = getsheet.cell(row=i+1, column=col_maker).value             # 原廠膠盒廠商
                T13  = getsheet.cell(row=i+2, column=col_num).value               #原廠端子料號
                HT25 = getsheet.cell(row=i+2, column=col_maker).value             #原廠端子廠商
                if T13 is None:
                    pass
                else:
                    j=1 #while迴圈迭代器
                    while T13 is not None:
                        if T13.startswith("T:"):
                            break
                        else:
                            T13  = getsheet.cell(row=i+2+j, column=col_num).value             #原廠端子料號
                            HT25 = getsheet.cell(row=i+2+j, column=col_maker).value           #原廠端子廠商
                        j += 1                
                # cscc_conn_name[接頭名稱]={ 膠盒料號, 廠商, 端子料號,廠商} 
                dit[L12]= [H13.lstrip('H:'),HF25,T13,HT25]
                theSameName(dit, L12, getsheet, i, 2,col_num,col_maker)

        else: # 功能名稱已在CSCC字典中,改為添加值
            L12=getsheet.cell(row=i, column=col_function).value.upper().strip().replace('_',' ').replace('-',' ')
            H13 = getsheet.cell(row=i + 1, column=col_num).value  # 原廠膠盒料號
            # 如果H13是空字符,那就將H13由原本指定的列數再往下一列,其餘元素跟著往下偏移
            if H13 is None:
                H13 = getsheet.cell(row=i + 2, column=col_num).value.lstrip('H:')  #原廠膠盒料號
                HF25 = getsheet.cell(row=i + 2, column=col_maker).value            #原廠膠盒廠商
                T13  = getsheet.cell(row=i+3, column=col_num).value                #原廠端子料號
                HT25 = getsheet.cell(row=i+3, column=col_maker).value              #原廠端子廠商
                if T13 is None:
                    pass
                else:
                    j=1 #while迴圈迭代器
                    #如果T13不是以T:開頭,且不是None,就再往下一行搜尋
                    while T13 is not None:
                        if T13.startswith("T:"):
                            break
                        else:
                            T13  = getsheet.cell(row=i+3+j, column=col_num).value             #原廠端子料號
                            HT25 = getsheet.cell(row=i+3+j, column=col_maker).value           #原廠端子廠商
                        j += 1
                # cscc_conn_name[接頭名稱]={ 膠盒料號, 廠商, 端子料號,廠商}
                dit[L12].append(H13)
                dit[L12].append(HF25)
                dit[L12].append(T13)
                dit[L12].append(HT25)
                print("The connector %s append %s" % (L12, (HF25+' '+H13)))         
                # 如果原行數下一行的值非"None",比對其是否以H:開頭,如果是就加入list中
                theSameName(dit, L12, getsheet, i, 3,col_num,col_maker)


            else:  # 如果H13不是None, 就不用往下偏移
                HF25 = getsheet.cell(row=i + 1, column=col_maker).value  # 原廠膠盒廠商
                T13  = getsheet.cell(row=i+2, column=col_num).value               #原廠端子料號
                HT25 = getsheet.cell(row=i+2, column=col_maker).value             #原廠端子廠商
                if T13 is None:
                    pass
                else:
                    j=1 #while迴圈迭代器
                    #如果T13不是以T:開頭,且不是None,就再往下一行搜尋
                    while T13 is not None:
                        if T13.startswith("T:"):
                            break
                        else:
                            T13  = getsheet.cell(row=i+2+j, column=col_num).value             #原廠端子料號
                            HT25 = getsheet.cell(row=i+2+j, column=col_maker).value           #原廠端子廠商
                        j += 1                                
                # cscc_conn_name[接頭名稱]={ 膠盒料號, 廠商, 端子料號,廠商}
                dit[L12].append(H13.lstrip('H:'))
                dit[L12].append(HF25)
                dit[L12].append(T13)
                dit[L12].append(HT25)
                print("The connector %s append %s" % (L12, (HF25+' '+H13)))
                theSameName(dit, L12, getsheet, i, 2,col_num,col_maker)
  
    return dit

#===========================以下為Connector比對專用函數=====================================
def cute(item):
    return item.upper().strip().replace('_', ' ').replace('-',' ')

def fuzzy(connector, dit, getsheet, i): #模糊比對
    global num
    print("[%s]The connector  %s is in fuzzy mode" % (i, connector))
    for csccname in dit:
        s1=set(connector.split())
        s2=set(csccname.split())
        #=========進入Find()比對模式========
        #主要解決HEAD_LAMP_RH與HEAD_LAMP會無法辨識的問題
        if len(connector) > len(csccname):   #如果connector長度比cscc_name大,就以前者為基準

            if connector.find(csccname) >=0: #有比對出cscc_name字串出現在connector中
                print("[+]the connector is in (find) mode")
                if len(dit[csccname]) >4:
                    moreConnector(csccname, dit, getsheet,i)
                    break
                else:
                    print("[+]The connector in (find) is match %s" % csccname)
                    getsheet.cell(row=i, column=25).value=dit[csccname][1]
                    getsheet.cell(row=i, column=26).value=dit[csccname][0].lstrip('H:')
                    getsheet.cell(row=i, column=30).value=dit[csccname][3]
                    getsheet.cell(row=i, column=31).value=dit[csccname][2]
                    if dit[csccname][1] != dit[csccname][3] and dit[csccname][3] is not None:
                        getcbsheet.cell(row=i, column=30).fill=Color_red
                    num += 1
                    break
                   
        elif csccname.find(connector) >=0:
            print("[+]the connector is in (find) mode")
            if len(dit[csccname]) >4:
                moreConnector(csccname, dit, getsheet,i)
                break                  
            else:
                print("[+]The connector in (find) is match %s" % csccname)
                getsheet.cell(row=i, column=25).value=dit[csccname][1]
                getsheet.cell(row=i, column=26).value=dit[csccname][0].lstrip('H:')
                getsheet.cell(row=i, column=30).value=dit[csccname][3]
                getsheet.cell(row=i, column=31).value=dit[csccname][2]
                if dit[csccname][1] != dit[csccname][3] and dit[csccname][3] is not None:
                    getcbsheet.cell(row=i, column=30).fill=Color_red
                num += 1
                break
        #==============進入set比對模式========== 
        #此模式主要可對應RH_HEAD_Lamp與HEAD_LAMP_RH會無法辨識的問題
        #差異的部份小於三處,就視為相同,且忽略功能名稱字串小於兩個的項目               
        elif len(s1.symmetric_difference(s2)) <3 and len(s1)>2 and len(s2)>2 : 
            print("[+]the connector is in (set) mode")
            if len(dit[csccname]) > 4:
                moreConnector(csccname, dit, getsheet,i)
                break             
            else:                
                print("[+]The connectoe in (set) is find match %s" % csccname)
                getsheet.cell(row=i, column=25).value=dit[csccname][1]
                getsheet.cell(row=i, column=26).value=dit[csccname][0].lstrip('H:')
                getsheet.cell(row=i, column=30).value=dit[csccname][3]
                getsheet.cell(row=i, column=31).value=dit[csccname][2]
                if dit[csccname][1] != dit[csccname][3] and dit[csccname][3] is not None:
                    getcbsheet.cell(row=i, column=30).fill=Color_red
                num +=1
                break
            
                
        else:
            getsheet.cell(row=i, column=26).value ="Can't match"
        

    
def moreConnector(connector, dit, getsheet, i):
    global num
    for j in range(0, len(dit[connector]),4):
        #如果cscc字典中有鍵的值與connector list的料號相同,就指定將該鍵的值填入connector list中
        if cute((dit[connector][j+1]+' '+dit[connector][j]))==cute(getsheet.cell(row=i, column=18).value):
            getsheet.cell(row=i, column=25).value=dit[connector][j+1]
            getsheet.cell(row=i, column=26).value=dit[connector][j].lstrip('H:')
            getsheet.cell(row=i, column=30).value=dit[connector][j+3]
            getsheet.cell(row=i, column=31).value=dit[connector][j+2]
            print("[%s]find connector %s match cscc,and is %s" % (i,connector,(dit[connector][j+1]+' '+dit[connector][j])) )
            if dit[connector][j+1] != dit[connector][j+3] and dit[connector][j+3] is not None:
                getcbsheet.cell(row=i, column=30).fill=Color_red
            num += 1
            break
        #否則就比對connector_list的接頭孔數(column=16(P))是否與cscc中的日產編號型式相同,是的話填入connector_list中(意指廠商使用LOCAL件)
        elif int(getsheet.cell(row=i, column=16).value) == int(re.findall(r'\w*(\d\d)\D*',dit[connector][j])[0]):
            getsheet.cell(row=i, column=25).value=dit[connector][j+1]
            getsheet.cell(row=i, column=26).value=dit[connector][j].lstrip('H:')
            getsheet.cell(row=i, column=30).value=dit[connector][j+3]
            getsheet.cell(row=i, column=31).value=dit[connector][j+2]
            print("[%s]find connector %s match 日產編號 cscc,and is %s" % (i,connector,(dit[connector][j+1]+' '+dit[connector][j])))
            if dit[connector][j+1] != dit[connector][j+3] and dit[connector][j+3] is not None:
                getcbsheet.cell(row=i, column=30).fill=Color_red
            num += 1
            break
        else:
            pass



#================載入樂榮CSCC ============
CSCC_Name=['24010-KN711-CSCC-161214.xlsx', '24012-KN711-CSCC-161214.xlsx',
           '24068-KN711-CSCC-161214.xlsx','24023-KN711-CSCC-170307.xlsx']
print("創建Excel物件....CSCC")
cscc_conn={}
for CsccName in CSCC_Name:
    print("載入 %s" % CsccName)
    wb, sheetname=loadExcel(CsccName)
    for shet in sheetname:
        getsheet=wb.get_sheet_by_name(shet)
        for i in range(25, int(getsheet.max_row)+1):
            cscc_conn=cscc(cscc_conn, i, getsheet,12,13,25)

#================載入矢崎CSCC==============
CsccName1='24011-KN711-CSCC-161104.xlsx'
print("載入 %s ..." % CsccName1)
wb3,sheetname3=loadExcel(CsccName1)
getsheet3=wb3.get_sheet_by_name(sheetname3[0])
for i in range(25, int(getsheet3.max_row)+1):
    cscc_conn=cscc(cscc_conn,i,getsheet3,14,15,27)

#===============比對Connector List與cscc內容=======
connector_list='GPKMC_Connector list_R-2版_NEW2.xlsx'
print("載入 %s" % connector_list)
wb2,sheet2=loadExcel(connector_list)
getcbsheet=wb2.get_sheet_by_name(sheet2[0])
for i in range(4, int(getcbsheet.max_row)+1):
    if getcbsheet.cell(row=i, column=10).value is None: #如果功能名稱是None,就跳過
        pass
    else:
        JS10=cute(getcbsheet.cell(row=i, column=10).value)
        # 如果 connector list的功能名稱有出現在cscc字典中
        if JS10 in cscc_conn:
            if len(cscc_conn[JS10]) > 4: #如果len >4 代表有複數顆接頭
                moreConnector(JS10, cscc_conn, getcbsheet,i)
            else: #len沒有大於3,代表接頭與料號為一對一的關係
                getcbsheet.cell(row=i, column=25).value=cscc_conn[JS10][1]
                getcbsheet.cell(row=i, column=26).value=cscc_conn[JS10][0].lstrip('H:')
                getcbsheet.cell(row=i, column=30).value=cscc_conn[JS10][3]
                getcbsheet.cell(row=i, column=31).value=cscc_conn[JS10][2]
                print("[%s]find connector %s match cscc" % (i,JS10))
                if cscc_conn[JS10][1] != cscc_conn[JS10][3] and cscc_conn[JS10][3] is not None:
                    getcbsheet.cell(row=i, column=30).fill=Color_red
                num += 1              

        else: #如果功能名稱不在cscc中,就進入模糊比對
            fuzzy(JS10,cscc_conn,getcbsheet,i)

print("\n")
print(str(datetime.now()))
print("Match %s items" % num)

print("Save File...")
wb2.save('GPKMC_Connector list_test7T.xlsx')
print("Done")
