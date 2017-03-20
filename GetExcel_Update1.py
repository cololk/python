#本程式作用為可比對出廠商自訂料號與接頭原廠編號的對應關係
import openpyxl
import re
import time

#=======載入材規資料=======
CSCC_Name=['24014-LB110-CSCC--C-170213.xlsx', '24015-LB110-CSCC--C-161031.xlsx',
                           '24068-LB110-CSCC--C-160923.xlsx','24093-LB110-CSCC--0-161031.xlsx']
print("創建Excel物件...載入CSCC")
cscc_no={}
for CsccName in CSCC_Name:
    print("載入 %s" % CsccName)
    wb = openpyxl.load_workbook(CsccName)
    sheetname = wb.get_sheet_names()
    getsheet = wb.get_sheet_by_name(sheetname[0])

    #創建材規的附件與料號對照字典   
    for i in range(27, int(getsheet.max_row)+1):           #因24014的材規從27列才開始是資料,上限需+1才能讀取到最後一列本身
        DN118 = getsheet.cell(row=i, column=118).value     #DN欄位轉換為數字為118
        DN118_str=str(DN118).strip()                       #全部轉換成字串型式,且去除頭尾字符
        L12 = getsheet.cell(row=i, column=12).value
        if L12 is None:
            pass
        else:
            L12strip=L12.lstrip('H:').lstrip('T:').strip()            #去除字頭的H:或者T: , 最後去除前後空白
        if DN118 is None:                                             #花了一堆時間確認noneType型式的判斷法...= =
            pass
        else:
            DN118strip=DN118_str.lstrip("附件")                 #去除字串頭的"附件"兩個字        
            #字典檔開始增加資料,格式為cscc_no[A]=[ B, C], 字典'鍵'對應的'值'為串列型式,        
            #cscc_no[A][0]為料件原廠編號, cscc_no[A][1]為料件廠商
            cscc_no[DN118strip] = [L12strip,  getsheet.cell(row=i, column=45).value]  



# ====載入CONNECTOR LIST========
print("載入Connector_List ...")
cb=openpyxl.load_workbook('LCH1_CONN_LIST_161213.xlsx')
cbsheetname= cb.get_sheet_names()
getcbsheet=cb.get_sheet_by_name(cbsheetname[0])

# 創建接頭料號與功能名稱對照字典
connector_list={}
# 正規表示法,共三種CASE
Reg1=re.compile(r'\d+-\d+-\d+')
Reg2=re.compile(r'\d+-\d+')
Reg3=re.compile(r'\w+')

for k in range( 4, int(getcbsheet.max_row)+1):
    J10=getcbsheet.cell(row=k , column=10).value  #J欄為接頭料號欄,對應數字為10
    if J10 is None or J10 ==' ':
        pass
    else:        
        R_J10=Reg1.findall(J10)
        if len(R_J10)==0:
            R_J10=Reg2.findall(J10)
            if len(R_J10)==0:
                R_J10=Reg3.findall(J10)
        R_J10_str=str(R_J10[0])
        print("Data Increase....row:%s" % k)
        #建立connector_list 字典 ,此為建立字典的第二種方式,利用function[A]=B
        connector_list[R_J10_str] = getcbsheet.cell(row=k, column=9).value

print("Done")
#------------ 以上為資料庫建立---------------------

  
# ======載入待比對資料EXCEL表=======
ob = openpyxl.load_workbook('LCH-DSHEET-170114.xlsx')
obsheetname = ob.get_sheet_names()
getobsheet=ob.get_sheet_by_name(obsheetname[1])

#創建待比對資料串列
object_no=[]
for j in range(3,int(getobsheet.max_row)+1):
    #增加進去的串列資料要加上str()方法,以確定都是以字串的形式加入,不能是數字,且去除頭尾字符
    object_no.append(str(getobsheet.cell(row=j, column=2).value).strip()) 


#=====找出待比對資料對應的料號((主程序))=======         
print("比對程序展開......")
time.sleep(1)
find=0  #找出的資料計數器初始值
# 找出的待比對廠商編號對應的原廠料號
final_list={}  
for attatch_no in object_no:
    flag=0
    for data in cscc_no:
        if data==attatch_no:
            print("廠商料號 %s 於CSCC中對應的原廠料號為 %s %s for %s" % (attatch_no, cscc_no[data][1],  cscc_no[data][0], data))            
            flag=1        #找到資料,旗標就設定為1 
            find=find+1
            # 產生待比對附件編號對照料號表
            final_list[attatch_no]=[cscc_no[data][0], cscc_no[data][1]]
        else:
            pass
    if flag==0:
        print("廠商料號 %s 無對應的原廠料號"% attatch_no)


print("比對完畢")
print("共有 %s 筆資料, 比對出 %s 項" % (len(object_no),find))
print(" ")


# =====找出附件料號對應原廠編號的接頭功能======
print("廠商料號對應的原廠料號以及功能為: ")
time.sleep(1)
find=0  #成功比對出的資料計數器
# 字典格式:  finalData_list [A]= [ 原廠料號, 廠商名稱, 接頭對應功能]

finalData_list={}
for attatch_no in final_list:
    for data1 in connector_list:
        if data1==final_list[attatch_no][0]:
            finalData_list[attatch_no]=[final_list[attatch_no][0], final_list[attatch_no][1],connector_list[data1]]
            print("附件編號 %s 對應的原廠料號為%s %s ,此接頭功能為 %s" % (attatch_no, finalData_list[attatch_no][1], finalData_list[attatch_no][0], finalData_list[attatch_no][2] ))
            find=find+1
        else:
            pass

print("比對出功能的料號共有%s 項" % find)
print("Done.")

#比對出功能的接頭會少於成功比對出廠商料號對應原廠料號的項目
#原因是比對出功能這個FUNCTION是靠原廠料號找出connector_list對應的功能
#如果廠商用代用件,就會找不出接頭對應的功能
    
   






