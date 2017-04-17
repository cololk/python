#-*- coding:utf-8 -*-
#適用python34
#本程序示範multiprocess在openpyxl的應用
#本程式裡用queue進行列隊寫入處理,writer()函數並無使用多進程,所以會依序寫入
import multiprocessing as mp
import queue, os, time
import random as rd
import openpyxl


class wsDiff(object):
    def __init__(self, row, column, value):
        self.row = row
        self.column = column
        self.value = value


def ws_job(wb, ws_idx):
    ws = wb.worksheets[ws_idx]
    print('[PID %s]: process (%s)' % (os.getpid(), ws.title))

    # *** DO SOME STUFF HERE***
    # Simulate workload
    time.sleep(rd.randrange(3, 6)) #模擬處理速度,讓程序暫停2~5秒,測試是否併行

    diff = []
    for i in range(1, 19):
        if ws.cell(row=i, column=1).value == 'TM3':
            print("Process [%s] Find TM3 on sheet %s ,and line %s" % (os.getpid(),ws_idx+1,i))
            #ws.cell(row=i, column=3).value = 'OK' ←注意!,並不在ws_job()函數中進行資料寫入,所以註解掉這行
            diff.append( wsDiff(i, 3, 'OK') ) #改而將要寫入的資料以及座標資料回傳

    return diff

#job()主要將傳入f_queue的index資訊,交給ws_job()處理,並將處理結果傳入q_queue
#由於f_queue傳入的是一連串依序傳入的index,所以任務會被列隊處理(一個傳入就處理一次,而不是一連串一起處理)
#補充說明,queue本身就是種列隊式的模塊函數
#注意job()主要功能為queue資料的管理, f_queue的輸入,經處理後輸出至w_queue,以及當f_queue已經沒資料要關閉進程
def job(fq, q, wb): 
    while True:
        try:
            ws_idx = fq.get_nowait() #queue加上.get_nowait()為非阻塞式,意即不是一個處理完才換下一個,而是能同時消化幾個就同時處理幾個
        # queue.Empty為當queue已是空,但卻又有get()想從中取物時出現, 這行只會當queue中的任務已經處理完才會跳到這個階段
        except queue.Empty: 
            print('PID %s: exit job' % os.getpid())
            exit(0) #結束進程,離開

        q.put((ws_job(wb, ws_idx), ws_idx))
        time.sleep(0.1)


def writer(q, wb): # q_queue中放的是要寫入sheet的資料串列
    print('Start Writer Process with PID %s' % os.getpid())
    while True:
        try:
            diff, i_ws = q.get()
        except ValueError:
            print('writer ValueError exit(1)')
            exit(1)

        if diff == None:
            print('[Writer Process PID]: %s save example_output.xlsx'% os.getpid())
            wb.save('example_output.xlsx')
            exit(0)

        ws = wb.worksheets[i_ws] #i_ws要跟著傳入q_queue的原因是因為,這是唯一可以識別要寫的資料要寫入哪個sheet的指示
        print('[Writer PID %s]: write sheet [%s] from diff' % (os.getpid(), ws.title))
        for d in diff:
            ws.cell(row=d.row, column=d.column).value = d.value


def mpRun():
    wb = openpyxl.load_workbook('example.xlsx')

    f_queue = mp.Queue()
    for i in range(len(wb.worksheets)):
        f_queue.put(i)

    w_queue = mp.Queue()
    w_process = mp.Process(target=writer, args=(w_queue, wb)) #在這一步驟時,尚無涉及f_queueueue內容,所以writer()僅是展開進入待命
    w_process.start()

    #需注意此處wb也有傳入多進程函數中,已進入複製階段
    pool = [mp.Process(target=job, args=(f_queue, w_queue, wb)) for p in range(os.cpu_count())]
    for p in pool:
        p.start()

    for p in pool:
        p.join()

    # Terminate Process w_process(主管寫入資料的進程程序) after all Sheets done
    w_queue.put((None, None))
    w_process.join()

    print('Done')

if __name__=='__main__':
    mpRun()