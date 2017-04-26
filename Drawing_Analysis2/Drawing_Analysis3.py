#-*- coding:utf-8 -*-
#本程式可進行圖譜的自動分析
#適用Python2.7版,原因為此程式使用wxPython
#但現已可適用Python36,因已於python36安裝了相容於此版本的wxpython
#具備開啟檔案話框功能
#具備多線程能力,可避免單線程時移動視窗會出現無回應的狀況
#test01-於part1加入multiprocess處理方式,其餘部分無法使用multiprocess,避免記憶體爆表當機
#4/26更新:追加防呆彈出式對話框,並更新檔名為Drawing_Analysis3.py
'''
2017/4/22紀錄:在python3(wx為phonex版本)運行時會出現AttributeError: module 'wx' has no attribute 'OPEN'的問題
因此修改語法,將參數wx.OPEN改為wx.FD_OPEN後,問題排除
2017/4/26紀錄:程式無法以pyinstaller打包,因為其不支援python3.6, 也無法使用Py2exe,因為其也不支援python3.6
而overflow上面推薦的cx_freeze也無法適用,因為聽說他的程式碼不能包含if __name__=='__main__'敘述句
'''
import wx
import sys
import os
import openpyxl
from openpyxl.styles import PatternFill
import frame2
import re
import threading
import time #使用計時模組
import multiprocessing as mp

Color_Green=PatternFill(fgColor='008000', fill_type="solid")
Color_Red=PatternFill(fgColor='ff0000', fill_type="solid")
Reg=re.compile(r'\d\d\d\d\d')
lis=[] # lis的組成內容=[24010,SPEC_CODE,24011,SPEC_CODE,24012,SPEC_CODE]

def GereratePartContent(num,getsheet):
        PartContent={}  # PartContent={'CM1':['AA1J','AA1B'], 'PD1':['AA1R','AB1A',...],...}
        for j in range(5, getsheet.max_row+1): # j是圖譜的列數
                if getsheet.cell(row=j, column=getsheet.max_column-(16-num)).value==u'\u25cf':
                        for k in range(11, getsheet.max_column-17): # k是車規的欄數
                                if getsheet.cell(row=j, column=k).value==u'\u25cf':
                                        if getsheet.cell(row=j, column=7).value not in PartContent: # 如果此SPEC CODE是第一次在這個件號出現,就設定其值
                                                PartContent[getsheet.cell(row=j,column=7).value]=[getsheet.cell(row=4,column=k).value]
                                                #print("Add %s"% getsheet.cell(row=4,column=k).value )
                                        else: #否則就改為添加其值
                                                PartContent[getsheet.cell(row=j,column=7).value].append(getsheet.cell(row=4,column=k).value)
                                                #print("Add %s"% getsheet.cell(row=4,column=k).value )
        return  getsheet.cell(row=4, column=getsheet.max_column-(16-num)).value, PartContent


#-------------多執行緒起手式(標準樣板)--------------#
myEVT_COUNT = wx.NewEventType()
EVT_COUNT = wx.PyEventBinder(myEVT_COUNT, 1)
class CountEvent(wx.PyCommandEvent):

    def __init__(self, etype, eid, value=None):
        wx.PyCommandEvent.__init__(self, etype, eid)
        self._value = value

    def GetValue(self):
        return self._value

#===============視窗介面框架================

#object要代入欲被轉換為Stdin輸出容器的TextCtrl物件
#此class主要處理標準輸出轉至視窗介面輸出的過程
class RedirectText(object):
    def __init__(self,aWxTextCtrl):
        self.out=aWxTextCtrl

    def write(self,string):
        self.out.WriteText(string)

class Myframe(frame2.MyFrame1):
        def init_Myframe(self): #自行定義的方法,設定初始化狀態欄位顯示的文字
                self.m_textCtrl1.SetValue(u"圖譜總檔位置")
                self.m_textCtrl3.SetValue(u"分頁檔案位置")

        def init_redirectText(self):  #這是自行定義的方法,在後面的主函數中要特別呼叫
                #輸出重新導向
                redir=RedirectText(self.m_textCtrl2)
                sys.stdout=redir

        def openfile(self,event): #可開啟檔案對話框
                global file
                wildcard="Excel(*.xlsx)|*.xlsx"
                dlg = wx.FileDialog(self, u"挑選你要的檔案_須轉檔為xlsx格式", os.getcwd(),"", wildcard, wx.FD_OPEN)
                dlg.ShowModal()
                file = dlg.GetPath() #從此處取得要開啟的檔案名稱與路徑
                dlg.Destroy()
                self.m_textCtrl1.SetValue(file)

        def OpenExcelSheet(self,event):
                global file2
                wildcard="Excel(*.xlsx)|*.xlsx"
                dlg = wx.FileDialog(self, u"挑選你要的檔案_須轉檔為xlsx格式", os.getcwd(),"", wildcard, wx.FD_OPEN)
                dlg.ShowModal()
                file2 = dlg.GetPath() #從此處取得要開啟的檔案名稱與路徑
                dlg.Destroy()
                self.m_textCtrl3.SetValue(file2)


        def loadExcel(self,event):  #採用多執行緒,參考Part1
                worker = loadExcel_Thread(self,1)
                worker.start()


        def loadExcelSheet(self,event):  #採用多執行緒,參考Part2
                worker1 = loadExcelSheet_Thread(self,1)
                worker1.start()
                                        

        def analysis(self,event):  #採用多執行緒,參考Part3
                worker2 = analysis_Thread(self,1)
                worker2.start()
                
#==============多執行緒 part1=========================
class loadExcel_Thread(threading.Thread):
    def __init__(self, parent, value):
        threading.Thread.__init__(self)
        self._parent = parent
        self._value = value

    def run(self):
        print("Program Starting...")
        print(u"載入 %s..." % os.path.basename(file))
        wb=openpyxl.load_workbook(file)
        sheetname=wb.get_sheet_names()
        getsheet=wb.get_sheet_by_name(sheetname[0])
        print(u"載入%s..." % sheetname[0])
        if sheetname[0] != '圖譜總表':
            print(u"檔案內容錯誤,讀取的檔案內容須為'圖譜總表'")
            #以下為彈出式對話框程式碼,共需三行
            dlg=wx.MessageDialog(None,u"檔案內容錯誤,讀取的檔案內容須為'圖譜總表'",u"資料錯誤提醒",wx.OK)
            result=dlg.ShowModal()
            dlg.Destroy()
        else:
            #件號串列資料庫建立
            print(u"開始建立件號組成資料...")
            pool=mp.Pool()
            result=[]
            for i in range(17): #range(i)此處控制哪幾個件號需納入分析
                    res=pool.apply_async(GereratePartContent,args=(i,getsheet))
                    result.append(res)
            pool.close()
            pool.join
            for item in result:
                    lis.append(item.get())
                    print(u"件號 %s SPEC CODE組成建立" % item.get()[0])
            print("Done")
        evt = CountEvent(myEVT_COUNT, -1, self._value)
        wx.PostEvent(self._parent, evt)

#==============多執行緒part2======================
class loadExcelSheet_Thread(threading.Thread):
    def __init__(self, parent, value):
        threading.Thread.__init__(self)
        self._parent = parent
        self._value = value

    def run(self):
                #載入各分頁Excel
                global sheetname2, wb2
                print(u"開始載入分頁檔案")
                print(u"載入 %s..." % os.path.basename(file2))
                wb2=openpyxl.load_workbook(file2)
                sheetname2=wb2.get_sheet_names()
                if not sheetname2[0].startswith("DRAWING"):
                    print(u"檔案內容錯誤,讀取的檔案內容須為'DRAWING_LIST'")
                    dlg=wx.MessageDialog(None,u"檔案內容錯誤,讀取的檔案內容須為'DRAWING_LIST'",u"資料錯誤提醒",wx.OK)
                    result=dlg.ShowModal()
                    dlg.Destroy()  
                else:                 
                    print(u"分頁檔案共有 %s 個sheet" % len(sheetname2))
                    for i in range(len(sheetname2)):
                            print("[%s] %s" % (i+1,sheetname2[i]))
                    print(u"載入完成")
                    evt = CountEvent(myEVT_COUNT, -1, self._value)
                    wx.PostEvent(self._parent, evt)

#=============多執行緒part3=================
class analysis_Thread(threading.Thread):
    def __init__(self, parent, value):

        threading.Thread.__init__(self)
        self._parent = parent
        self._value = value

    def run(self):
                #展開比對程序 
                starttime=time.time()   
                print(u"展開比對程序...")
                for n in range(11, wb2.get_sheet_by_name(sheetname2[0]).max_column):
                        if wb2.get_sheet_by_name(sheetname2[0]).cell(row=3, column=n).value==u'件號':
                                g=n+1
                                print(u"取得分析起始欄位%s" % g)
                for h in range(0,len(sheetname2)):  #取得載入的EXCEL共有幾個Sheet
                        get=wb2.get_sheet_by_name(sheetname2[h])
                        print(u"載入%s"% sheetname2[h])
                        if str(Reg.findall(sheetname2[h])[0]) in str(lis[h][0]):  #須轉換為long型態才能跟sheetname2的頁面名稱匹配
                                print(u"取得 %s 對應圖譜總表資料 %s" % (sheetname2[h],lis[h][0]))
                                for i in range (5, get.max_row+1):
                                        if get.cell(row=i, column=7).value in lis[h][1]: #如果SPEC_CODE有出現在lis[1]中(這邊是指24010)
                                                for j in range(len(lis[h][1][get.cell(row=i, column=7).value])): #找出該件號該Spec_Code共出現幾個車規                   
                                                        for k in range(g, get.max_column+1):        
                                                                                                        # 如果依序遍歷車規發現有與該件號該SPE_CODE相同的就打上記號並標註為綠色
                                                                if get.cell(row=4, column=k).value==lis[h][1][get.cell(row=i, column=7).value][j]:
                                                                        if get.cell(row=i, column=k).value is not None: #如果該欄位不是空,那就注入綠色(代表該選有選)
                                                                                get.cell(row=i, column=k).value=u'\u25cf'
                                                                                get.cell(row=i, column=k).fill=Color_Green
                                                                        else: #否則該欄位就是空,那就注入紅色(代表該選沒選)
                                                                                get.cell(row=i, column=k).value=u'\u25cf'
                                                                                get.cell(row=i, column=k).fill=Color_Red

                path, filename = os.path.split(file2)
                saveFileName=path+'\\OUTPUT_'+filename                                          
                wb2.save(saveFileName)
                print(u"檔案產出: %s" % os.path.basename(saveFileName))
                endtime=time.time()
                total=endtime-starttime
                print("used time:", total)
                print("Done")
                evt = CountEvent(myEVT_COUNT, -1, self._value)
                wx.PostEvent(self._parent, evt)


if __name__=='__main__':
        app=wx.App()
        main_win=Myframe(None)
        main_win.init_Myframe()
        main_win.init_redirectText()
        main_win.Show()
        app.MainLoop()


        
    
