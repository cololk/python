#-*- coding:utf-8 -*-
import wx
import guiframe
import openpyxl
from openpyxl.styles import PatternFill

#=============設定儲存格顏色==================
Color_red=PatternFill(fgColor='DC143C', fill_type="solid")  #作用為如果廠商不同廠,可把儲存格顏色反紅
Color_green=PatternFill(fgColor='98fb98', fill_type="solid")

def trans(cgetsheet):
	lis={}
	for i in range(3,cgetsheet.max_row+1):
		lis[cgetsheet.cell(row=i,column=1).value]=cgetsheet.cell(row=i, column=2).value
	return lis
#===============載入Excel========================  
wb=openpyxl.load_workbook('INPUT.xlsx')
sheetname=wb.get_sheet_names()
for i in range(0,len(sheetname)):
    getsheet=wb.get_sheet_by_name(sheetname[i])
    if i==0:
    	L1=trans(getsheet)
    	s1=set(L1)
    else:
    	L2=trans(getsheet)
    	s2=set(L2)
def scan():    	
        diff=list(s1.symmetric_difference(s2))
        inter=list(s1.intersection(s2))

        #=============建立OUTPUT檔案格式=============
        ws=wb.create_sheet()
        ws.title='OUTPUT'
        ws.column_dimensions["A"].width=12
        ws.column_dimensions["D"].width=12
        ws.column_dimensions["B"].width=18
        ws.column_dimensions["E"].width=18
        ws.cell(row=1,column=1).value=sheetname[0]
        ws.cell(row=2,column=1).value='SPEC_NO'
        ws.cell(row=2,column=1).fill=Color_green
        ws.cell(row=2,column=2).value='FUNCTION'
        ws.cell(row=2,column=2).fill=Color_green

        ws.cell(row=1,column=4).value=sheetname[1]
        ws.cell(row=2,column=4).value='SPEC_NO'
        ws.cell(row=2,column=4).fill=Color_green
        ws.cell(row=2,column=5).value='FUNCTION'
        ws.cell(row=2,column=5).fill=Color_green

        #=========先將比對相同的部分填入OUTPUT===========
        for i in range (0, len(inter)):
                j=i+3  #轉換成Excel輸入欄位
                ws.cell(row=j, column=1).value=inter[i]
                ws.cell(row=j, column=2).value=L1[inter[i]]
                ws.cell(row=j, column=4).value=inter[i]
                ws.cell(row=j, column=5).value=L2[inter[i]]

        #========再將不同的部分填入====================
        k=1
        m=1
        for dif in diff:	
                if dif in L1:
                        ws.cell(row=j+k, column=1).value=dif
                        ws.cell(row=j+k, column=1).fill=Color_red
                        ws.cell(row=j+k, column=2).value=L1[dif]
                        k += 1
                else:
                        ws.cell(row=j+m, column=4).value=dif
                        ws.cell(row=j+m, column=4).fill=Color_red
                        ws.cell(row=j+m, column=5).value=L2[dif]
                        m += 1
        wb.save("OUTPUT.xlsx")
        

#=======視窗介面=============

class Myframe(guiframe.baseFormWind):

    def init_main_window(self):
        self.m_textCtrl1.SetValue(sheetname[0])
        self.m_textCtrl2.SetValue(sheetname[1])
    def analysis(self,event):
        scan()
        self.m_staticText4.SetLabel("Generated file----> OUTPUT.xlsx")


if __name__=='__main__':
    app=wx.App()
    main_win=Myframe(None)
    main_win.init_main_window()
    main_win.Show()
    app.MainLoop()

    
    

