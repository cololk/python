#-*- coding:utf-8 -*-
#本程式可比對兩個件件號之間SPEC CODE的差異

import openpyxl
lis=[[],[],[],[]]
j=0
wb=openpyxl.load_workbook('C:\\Python27\\TEST001.xlsx')
sheetname=wb.get_sheet_names()

for sheet in sheetname:
    
    getsheet=wb.get_sheet_by_name(sheet)
    for i in range(2,getsheet.max_row+1):
        lis[j].append(getsheet.cell(row=i, column=1).value)
    
    j += 1
        
s1=set(lis[0])
s2=set(lis[1])
s3=set(lis[2])
s4=set(lis[3])

print(s1.symmetric_difference(s2))
print(s3.symmetric_difference(s4))
